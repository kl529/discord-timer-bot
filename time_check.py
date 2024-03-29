import openpyxl as op
import time
import math
from datetime import datetime
from pytz import timezone

GOAL = {
    'JW': 10,
    'SM' : 10,
    'KO' : 10,
    'HM' : 10
}

# 1. 현재 시간 체크
# 파일 읽어서, 본인 이름에 대한 시간들을 모두 더해서 보여주기 + 일주일 목표 + 최종 목표
# 결과값 - 오늘 공부시간 / 이번주 공부시간 / 이번주 공부 목표 // 총 시간 / 최종 총 시간
def check_status(ctx):
    wb = op.load_workbook(r"result.xlsx") #Workbook 객체 생성
    ws = wb.active #활성화 된 시트 객체 생성

    total_time = 0
    today_time = 0
    week_time = 0

    semester = math.ceil(time.localtime().tm_mon / 3.0)

    for row in ws.rows:
        if ctx.message.author.name == row[1].value and semester == row[0].value and row[5].value is not None:
            total_time += row[5].value
            if row[2].value == time.strftime('%Y%m%d', time.localtime()):
                today_time += row[5].value

    return total_time, today_time, week_time


# 2. 시간 체크 시작하는 함수
def start_timer(ctx):

    semester = math.ceil(time.localtime().tm_mon / 3.0) #학기 체크
    author_name = ctx.message.author.name #이름
    now = datetime.now(timezone('Asia/Seoul'))
    date = now.strftime('%Y%m%d') # 날짜
    start_time = now.strftime('%Y-%m-%d %H:%M:%S') #시간

    wb = op.load_workbook(r"result.xlsx") #Workbook 객체 생성
    ws = wb.active #활성화 된 시트 객체 생성

    ws.append([semester, author_name, date, start_time]) # 열 추가
    wb.save(r"result.xlsx")

    return start_time


# 3. 시간 체크 종료하는 함수
def end_timer(ctx):

    wb = op.load_workbook(r"result.xlsx") #Workbook 객체 생성
    ws = wb.active #활성화 된 시트 객체 생성

    end_time = datetime.now(timezone('Asia/Seoul')).strftime('%Y-%m-%d %H:%M:%S') # 종료 시간

    for row in ws.rows:
        if row[1].value == ctx.message.author.name and row[4].value is None: # 시작을 했고, 닉네임이 같은 유저가 있으면
            gap = (datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S') - datetime.strptime(row[3].value, '%Y-%m-%d %H:%M:%S')).seconds
            row[4].value = end_time #종료시간을 넣어줌.
            row[5].value = gap #얼마나 공부했는지 넣어줌

            wb.save(r"result.xlsx")

            return end_time, time_stamp_to_time(gap)

    return 0,0


#초기 엑셀 세팅 -> 필요 없음.
def setting():
    wb = op.load_workbook(r"result.xlsx") #Workbook 객체 생성
    ws = wb.active #활성화 된 시트 객체 생성 

    col_names = ['학기', '이름', '날짜', '시작시간', '종료시간', '공부시간']
    for seq, name in enumerate(col_names):
        ws.cell(row=1, column=seq+1, value=name)

    wb.save(r"result.xlsx") #결과 엑셀파일 저장


# second를 시간 / 분 / 초로 변환하는 함수
def time_stamp_to_time(ts, seperate=False):
    hour = ts // 3600
    ts -= hour*3600
    minute = ts // 60
    second = ts - minute*60 

    if seperate:
        return (int(hour), int(minute), int(second))

    return_string = ''
    return_string += f'{int(hour)}시간 ' if hour != 0 else ''
    return_string += f'{int(minute)}분 ' if minute != 0 else ''
    return_string += f'{int(second)}초 ' if second != 0 else ''

    return f'{int(hour)}시간 {int(minute)}분 {int(second)}초'


# 학기가 끝날때, 모든 공부시간을 정리하는 함수
def calculate_timer():
    wb = op.load_workbook(r"result.xlsx") #Workbook 객체 생성
    ws = wb.active #활성화 된 시트 객체 생성

    semester = math.ceil(time.localtime().tm_mon / 3.0) #학기 체크

    result_dic = {'Karice' : 0, 'god_life': 0, 'kwanok': 0, '김선만': 0}

    for row in ws.rows:
        semester_input = row[0].value # 학기
        nickname_input = row[1].value # 닉네임
        date_input = row[2].value #시작 날짜
        start_time_input = row[3].value #시작 시간
        end_time_input = row[4].value #종료시간
        study_time = row[5].value #공부 시간

        if semester_input == semester and study_time:
            result_dic[nickname_input] += study_time
            
    result_dic = dict(sorted(result_dic.items(), key=(lambda x:x[1]), reverse=True)) # 정렬
    result_dic = [{'name' : item, 'time' : time_stamp_to_time(result_dic[item]), 'pass' : ('🎉' if result_dic[item] >= 360000 else '❌')} for item in result_dic] # second를 시간으로 변경

    message = f"{semester} 학기도 모두 고생하셨습니다. \n \n\
🥇 1등 - {result_dic[0]['name']} : {result_dic[0]['time']} ( {result_dic[0]['pass']} ) \n\n \
🥈 2등 - {result_dic[1]['name']} : {result_dic[1]['time']} ( {result_dic[1]['pass']} ) \n\n \
🥉 3등 - {result_dic[2]['name']} : {result_dic[2]['time']} ( {result_dic[2]['pass']} ) \n\n \
🏆 4등 - {result_dic[3]['name']} : {result_dic[3]['time']} ( {result_dic[3]['pass']} )"

    return message
